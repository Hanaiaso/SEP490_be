"""
Sinh deliverable cua dot kiem thu L4 (Report_5_4 v1.3) tu ket qua chay THAT cua Playwright.

  1. Dien Status / Defect ID / Notes vao Report_5_4_L4-E2ETests_VietTien_v1_3.xlsx
     (6 sheet chi tiet; sheet "TestCase List" tu cap nhat vi cot G/H la CONG THUC
      tro sang sheet chi tiet, va o tong hop hang 3-4 la COUNTIF).
  2. tests/L4_KET_QUA_TOM_TAT.md   <- so lieu tong hop de trich vao slide
  3. tests/L4_ROUTE_DRIFT.md       <- lech route/endpoint giua workbook va code that

Nguon du lieu:
  - SEP490_fe/frontend/e2e/playwright-report/results.json

Quy uoc phan loai (quan trong nhat cua bao cao nay):
  - Pass    : chay duoc va dung ky vong workbook.
  - Fail    : co man hinh / co endpoint nhung hanh vi SAI so voi workbook.
  - Blocked : khong co man hinh / khong co endpoint de kiem (module chua trien khai,
              hoac workbook ghi route khong ton tai). Day KHONG phai loi khi chay test.

Chay:  python tools/l4_report.py
"""
import json
import os
import re
from collections import OrderedDict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "Report_5_4_L4-E2ETests_VietTien_v1_3.xlsx")
RESULTS = os.path.join(
    os.path.dirname(ROOT), "SEP490_fe", "frontend", "e2e",
    "playwright-report", "results.json")

SHEETS = [
    "L4-CriticalPaths", "L4-UserJourneys", "L4-Permissions",
    "L4-SessionManagement", "L4-Responsive (Optional)", "L4-AdminMarketing",
]
COT_STATUS, COT_DEFECT, COT_NOTES = 12, 13, 14   # L, M, N

# Case do vi KHONG CO man hinh / endpoint de kiem -> Blocked, khong phai Fail.
# Moi dong: test id -> (defect id, ly do ngan gon dua vao cot Notes)
BLOCKED = {
    "L4-CP-05":  ("DEF-L4-002", "GET /api/delivery/trips -> 404. Module chuyen giao hang chua trien khai (trung DEF-L3-004)."),
    "L4-UJ-06":  ("DEF-L4-002", "GET /api/delivery/collections -> 404. Thu COD chua co endpoint (trung DEF-L3-004)."),
    "L4-UJ-11":  ("DEF-L4-003", "GET /api/inventory/count-sessions -> 404. Phien kiem ke chua trien khai (trung DEF-L3-005)."),
    "L4-UJ-13":  ("DEF-L4-003", "GET /api/inventory/low-stock-alerts -> 404. Canh bao ton thap chua co endpoint (trung DEF-L3-005)."),
    "L4-AM-10":  ("DEF-L4-003", "POST /api/marketing-posts/{id}/media -> 404. Upload media chua trien khai; khong kiem duoc NFR-SEC07."),
    "L4-AM-11":  ("DEF-L4-003", "GET /api/marketing-posts/{id}/metrics -> 404. Chi so tuong tac chua trien khai."),
    "L4-AM-05":  ("DEF-L4-004", "Workbook ghi /admin/ai-marketing nhung route do la stub ComingSoon (AdminPortal.tsx:205-207). Chuc nang that o /sales/ai-content-studio."),
    "L4-AM-07":  ("DEF-L4-004", "Workbook ghi /sales/pickup-arrangement nhung route khong ton tai trong SalesPortal.tsx (roi vao catch-all ve Dashboard)."),
}

# Ghi chu them cho cac case Pass dang chu y.
GHI_CHU_THEM = {
    "L4-CP-02": "Webhook that la POST /api/webhooks/sepay-callback + header x-sepay-token (workbook ghi /api/webhooks/sepay + HMAC).",
    "L4-CP-06": "Chay tren 10 san pham Active (seed bo sung 3 SP vi DB chi co 7).",
    "L4-PM-07": "API tra 403/404 dung; UI bao 'Khong the tai chi tiet don hang', khong lo ma don.",
    "L4-UJ-02": "Moc that: >=10tr co chiet khau bac; >=100tr chuyen sang 'Yeu cau bao gia' (bang DiscountTiers het bac o 100tr).",
    "L4-UJ-09": "POST /api/inventory/post-from-po -> 405 (khong mo duong tang ton thang tu PO) = dung BR-014.",
    "L4-SM-01": "Chay lai nhieu lan co the do: refresh token XOAY VONG moi lan refresh va authService.fetchWithToken khong co mutex -> nhieu request song song cung refresh se lam mot so request dung token cu. Xem DEF-L4-006.",
}

TEN_SHEET_THEO_MA = {
    "CP": "L4-CriticalPaths", "UJ": "L4-UserJourneys", "PM": "L4-Permissions",
    "SM": "L4-SessionManagement", "RS": "L4-Responsive (Optional)", "AM": "L4-AdminMarketing",
}


def doc_ket_qua():
    """Tra ve {test_id: ('passed'|'failed'|'skipped', thong_diep_loi_dau_tien)}."""
    with open(RESULTS, encoding="utf-8") as fh:
        data = json.load(fh)

    ket_qua = {}

    def duyet(suite):
        for spec in suite.get("specs", []):
            m = re.match(r"(L4-[A-Z]{2}-\d+)", spec.get("title", ""))
            if not m:
                continue
            tid = m.group(1)
            trang_thai, loi = "skipped", ""
            for t in spec.get("tests", []):
                for r in t.get("results", []):
                    st = r.get("status")
                    if st == "passed" and trang_thai != "failed":
                        trang_thai = "passed"
                    elif st in ("failed", "timedOut"):
                        trang_thai = "failed"
                        err = (r.get("error") or {}).get("message", "")
                        loi = re.sub(r"\x1b\[[0-9;]*m", "", err).strip().split("\n")[0][:300]
            # Mot Test ID chi can 1 lan do la ca case Fail.
            if tid in ket_qua and ket_qua[tid][0] == "failed":
                continue
            ket_qua[tid] = (trang_thai, loi)
        for con in suite.get("suites", []):
            duyet(con)

    for s in data.get("suites", []):
        duyet(s)
    return ket_qua


def phan_loai(tid, trang_thai, loi):
    """-> (status_ghi_vao_excel, defect_id, notes)"""
    if trang_thai == "passed":
        return "Pass", "", GHI_CHU_THEM.get(tid, "")
    if trang_thai == "skipped":
        return "Blocked", "", "Khong du tien dieu kien de chay."
    if tid in BLOCKED:
        defect, ly_do = BLOCKED[tid]
        return "Blocked", defect, ly_do
    # Voi case Fail da phan tich duoc nguyen nhan goc, ghi ket luan thay vi chuoi loi tho.
    return "Fail", DEFECT_THEO_CASE.get(tid, ""), GHI_CHU_FAIL.get(tid, loi)


# Ket luan da phan tich cho tung case Fail (thay cho thong diep assert tho).
GHI_CHU_FAIL = {
    "L4-CP-04": "sales-confirm tra 403 cho tai khoan SalesStaff. HandoverController.cs:35-36 khai "
                "role \"Sales\" — gia tri khong ton tai trong enum SystemRole -> khong the du 2 chu ky.",
    "L4-PM-04": "API chan Admin dung (ceo-decision 403, inventory adjust 403) NHUNG UI van cho Admin "
                "mo /warehouse/inv-management/stock-adjustment va /ceo/* (App.tsx:76,79).",
    "L4-SM-01": "KHONG on dinh — do 4/5 lan chay. Refresh token xoay vong moi lan refresh; "
                "authService.fetchWithToken khong co mutex nen request cham hon dung token da bi xoay "
                "-> 401 -> xoa phien va chuyen ve /login.",
    "L4-SM-05": "Sau khi day Carts.UpdatedAt lui 24:00:01, gio van cho thanh toan; khong co canh bao "
                "gia het han va khong bat lam moi gia. Khong tim thay logic 24h o CartService lan Cart.jsx.",
    "L4-UJ-14": "Dia chi da dung cho don hang bi xoa cung khoi DB. Don khong hong vi "
                "Orders.ShippingAddress la snapshot chuoi.",
}


# Defect that (co man hinh/endpoint nhung hanh vi sai).
DEFECT_THEO_CASE = {
    "L4-CP-04": "DEF-L4-001",
    "L4-PM-04": "DEF-L4-005",
    "L4-SM-05": "DEF-L4-007",
    "L4-SM-01": "DEF-L4-006",
    "L4-UJ-14": "DEF-L4-008",
}

DEFECTS = OrderedDict([
    ("DEF-L4-001", ("P1", "Ban giao 2 chu ky KHONG THE hoan tat: HandoverController.cs:35-36 khai "
                    "[Authorize(Roles=\"Sales\")] cho sales-confirm, nhung enum SystemRole "
                    "(Models/User.cs:60-70) khong co gia tri \"Sales\" (vai tro that la SalesStaff). "
                    "Moi tai khoan Sales deu bi 403 -> BR-034 khong bao gio du 2 chu ky -> khong xuat kho duoc. "
                    "Cung loi nay con o \"WarehouseManager\", \"SaleStaff\", \"SaleManager\" trong nhieu controller khac.")),
    ("DEF-L4-002", ("P1", "Module Delivery Trip / POD / thu COD chua trien khai: /api/delivery/trips, "
                    "/api/delivery/collections deu 404. Trung voi DEF-L3-004.")),
    ("DEF-L4-003", ("P2", "Thieu chuc nang: phien kiem ke (count-sessions), canh bao ton thap "
                    "(low-stock-alerts), upload media va chi so bai marketing. Trung voi DEF-L3-005.")),
    ("DEF-L4-004", ("P2", "Workbook ghi sai route: /admin/ai-marketing va /admin/marketing-history la "
                    "stub ComingSoon (AdminPortal.tsx:205-207); /sales/pickup-arrangement khong ton tai. "
                    "Chuc nang marketing that nam o /sales/ai-content-studio va "
                    "/sales-manager/ai-marketing-approval.")),
    ("DEF-L4-005", ("P1", "Phan quyen UI khong khop API: API chan Admin phe duyet nghiep vu dung "
                    "(ceo-decision va inventory adjust deu 403), NHUNG ProtectedRoute (App.tsx:76,79) "
                    "cho Admin mo ca /ceo/* va /warehouse/* -> Admin van vao duoc man duyet bao gia va "
                    "dieu chinh ton kho, bam nut roi moi bao loi. Vi pham NFR-SEC03/FT-09 NAC-02.")),
    ("DEF-L4-006", ("P2", "Khong co mutex khi refresh token: POST /api/auth/refresh-token XOAY VONG "
                    "refresh token (token cu lap tuc 401 — da do bang probe). "
                    "authService.fetchWithToken (src/services/authService.js:65-91) khong dong bo, nen khi "
                    "mot trang ban nhieu request can auth cung luc va access token het han, cac request "
                    "cham hon dung token da bi xoay -> 401 -> xoa phien va chuyen ve /login. "
                    "Phu: dong 85 xoa khoa 'user' trong khi AuthContext luu o 'authUser'.")),
    ("DEF-L4-007", ("P2", "Khong co co che het han snapshot gia gio hang 24 gio (BR-025): sau khi day "
                    "Carts.UpdatedAt lui 24:00:01, gio van cho thanh toan binh thuong, khong canh bao "
                    "'gia da het han' va khong bat lam moi gia.")),
])

# Lech route/endpoint giua workbook va code that — phat hien trong luc chay.
ROUTE_DRIFT = [
    ("L4-CP-02", "POST /api/webhooks/sepay + HMAC", "POST /api/webhooks/sepay-callback + header x-sepay-token", "Khac ten va khac cach xac thuc"),
    ("L4-CP-04", "POST /api/handover-records/{id}/confirm", "warehouse-confirm + sales-confirm", "Tach thanh 2 endpoint dual-confirm"),
    ("L4-CP-05", "/sales/delivery/arrangement + /api/delivery/trips", "Man hinh CO, endpoint KHONG", "Module Delivery Trip chua trien khai"),
    ("L4-UJ-03", "CEOPortal tab 'price-negotiation'", "/ceo (tab noi bo, khong co URL rieng)", "CEOPortal dung state tab, khong dieu huong bang URL"),
    ("L4-UJ-06", "/api/delivery/collections", "(khong ton tai)", "Thu COD chua co endpoint"),
    ("L4-UJ-08", "/warehouse/inv-management/quarantine", "/warehouse/inv-management/quarantine", "Trung khop (con co bi danh /warehouse/quarantine)"),
    ("L4-UJ-11", "/api/inventory/count-sessions", "(khong ton tai)", "Phien kiem ke chua trien khai"),
    ("L4-UJ-13", "/api/inventory/low-stock-alerts", "(khong ton tai)", "Canh bao ton thap di qua Notifications"),
    ("L4-AM-05", "/admin/ai-marketing", "/sales/ai-content-studio", "Route workbook la stub ComingSoon (AdminPortal.tsx:205-207)"),
    ("L4-AM-07", "/sales/pickup-arrangement", "(khong ton tai)", "Roi vao catch-all ve SalesDashboard (SalesPortal.tsx:425)"),
    ("L4-AM-09", "/admin/ai-marketing (duyet)", "/sales-manager/ai-marketing-approval", "Duyet bai nam o portal Sales Manager"),
    ("L4-AM-10", "POST /api/marketing-posts/{id}/media", "(khong ton tai)", "Upload media chua trien khai"),
    ("L4-AM-11", "/admin/marketing-history + /metrics", "stub ComingSoon; endpoint metrics khong ton tai", "Chi so tuong tac chua trien khai"),
]

# Ten dinh danh trong workbook -> du lieu that trong DB.
ANH_XA_DU_LIEU = [
    ("customer1@viettien.test", "customer.test@viettien.com", "Tai khoan seed, mat khau 123456"),
    ("warehouse1@viettien.test", "warehousestaff.test@viettien.com", "Gan kho WH-DEFAULT"),
    ("SKU-001", "WRAP-BB-1M2", "250.000d, ton 10.000 tai WH-DEFAULT — chon vi de dat moc 10tr/100tr"),
    ("SKU-002", "TOOL-CUT-5F", "25.000d"),
    ("WH-PROD", "WH-PE", "KHONG co kho ten WH-PROD; he thong co WH-DEFAULT / WH-TRADE / WH-PE"),
    ("SUP-01", "SUP-01 (do seed tao)", "Bang Suppliers ban dau rong"),
    ("ORD-SMOKE", "VT<yyyyMMddHHmmssfff>", "Bang Orders ban dau rong; moi don deu do test tao"),
]


def dien_workbook(ket_qua):
    wb = openpyxl.load_workbook(WORKBOOK)
    thong_ke = OrderedDict()
    chua_thay = []

    for ten in SHEETS:
        ws = wb[ten]
        for row in range(5, ws.max_row + 1):
            tid = ws.cell(row, 1).value
            if not isinstance(tid, str) or not re.match(r"^L4-[A-Z]{2}-\d+$", tid.strip()):
                continue
            tid = tid.strip()
            if tid not in ket_qua:
                chua_thay.append(tid)
                continue
            trang_thai, loi = ket_qua[tid]
            status, defect, notes = phan_loai(tid, trang_thai, loi)
            ws.cell(row, COT_STATUS).value = status
            ws.cell(row, COT_DEFECT).value = defect or None
            if notes:
                ws.cell(row, COT_NOTES).value = notes
            thong_ke[tid] = (ten, status, defect)

    wb.save(WORKBOOK)
    return thong_ke, chua_thay


def viet_tom_tat(thong_ke):
    tong = len(thong_ke)
    dem = {"Pass": 0, "Fail": 0, "Blocked": 0}
    for _, st, _ in thong_ke.values():
        dem[st] = dem.get(st, 0) + 1

    theo_sheet = OrderedDict()
    for tid, (sheet, st, _) in thong_ke.items():
        theo_sheet.setdefault(sheet, {"Pass": 0, "Fail": 0, "Blocked": 0})
        theo_sheet[sheet][st] += 1

    d = []
    d.append("# L4 E2E Test — ket qua tong hop (2026-08-12)\n")
    d.append(f"Tong so case trong workbook: **{tong}**\n")
    d.append("| Trang thai | So case | Ty le |")
    d.append("|---|---:|---:|")
    for k in ("Pass", "Fail", "Blocked"):
        d.append(f"| {k} | {dem[k]} | {dem[k] * 100.0 / tong:.1f}% |")

    d.append("\n## Cong cu da dung\n")
    d.append("| Cong cu | Pham vi | Ket qua |")
    d.append("|---|---|---|")
    d.append("| Playwright 1.x + Chromium that | 47 case E2E tren 2 viewport (1280x800 va 375x812) | "
             f"{dem['Pass']} xanh / {dem['Fail']} do / {dem['Blocked']} bi chan |")
    d.append("| sqlcmd (SQL Server local) | Doi chieu side-effect va dung trang thai khong tao duoc qua API | Xem L4_mutations_2026-08-12.sql |")
    d.append("| API that (khong mock) | Seed tien dieu kien + kiem tang API cho cac case Permission | — |")

    d.append("\n## Ket qua theo sheet\n")
    d.append("| Sheet | Pass | Fail | Blocked |")
    d.append("|---|---:|---:|---:|")
    for sheet, v in theo_sheet.items():
        d.append(f"| {sheet} | {v['Pass']} | {v['Fail']} | {v['Blocked']} |")

    d.append("\n## Defect\n")
    d.append("| Defect ID | Muc | Tom tat | Case lien quan |")
    d.append("|---|---|---|---|")
    for did, (muc, mo_ta) in DEFECTS.items():
        cases = sorted(t for t, (_, _, dfid) in thong_ke.items() if dfid == did)
        d.append(f"| {did} | {muc} | {mo_ta} | {', '.join(cases) or '—'} |")

    d.append("\n## Moi truong chay\n")
    d.append("- Backend: ASP.NET Core 8, `http://localhost:5050`, `ASPNETCORE_ENVIRONMENT=Development`.")
    d.append("- Database: **SQL Server local `VietTien22_L3`** (KHONG phai Azure — firewall Azure SQL "
             "chan IP may chay test, xem L4_HUONG_DAN_THUYET_TRINH.md).")
    d.append("- Frontend: Vite dev server `http://localhost:5173`, proxy `/api` + `/hubs` -> `127.0.0.1:5050`.")
    d.append("- Tai khoan: 11 tai khoan seed `*.test@viettien.com`, mat khau `123456`.")

    out = os.path.join(ROOT, "tests", "L4_KET_QUA_TOM_TAT.md")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(d) + "\n")
    return out, dem


def viet_route_drift():
    d = ["# L4 — lech giua workbook va code that\n",
         "Nguon: `Report_5_4_L4-E2ETests_VietTien_v1_3.xlsx` doi chieu voi code ngay 2026-08-12,",
         "phat hien trong luc chay Playwright chu khong phai doc code thuan tuy.\n",
         "## 1. Lech route / endpoint\n",
         "| Test ID | Workbook ghi | Thuc te trong code | Ghi chu |", "|---|---|---|---|"]
    for tid, wb_, that, note in ROUTE_DRIFT:
        d.append(f"| {tid} | `{wb_}` | `{that}` | {note} |")

    d.append("\n## 2. Lech ten du lieu\n")
    d.append("| Workbook ghi | Du lieu that | Ghi chu |")
    d.append("|---|---|---|")
    for a, b, c in ANH_XA_DU_LIEU:
        d.append(f"| `{a}` | `{b}` | {c} |")

    d.append("\n## 3. Vai tro khai trong controller nhung KHONG co trong enum SystemRole\n")
    d.append("`SystemRole` (Models/User.cs:60-70) chi co: Guest, Customer, SalesStaff, SalesManager, "
             "WarehouseStaff, AccountingStaff, CEO, Admin.\n")
    d.append("| Gia tri khai trong `[Authorize(Roles=...)]` | So cho xuat hien | Hau qua |")
    d.append("|---|---:|---|")
    d.append("| `Sales` | 2 (HandoverController) | **sales-confirm khong ai goi duoc -> BR-034 be gay** (DEF-L4-001) |")
    d.append("| `WarehouseManager` | 13 | Thua, khong cap quyen cho ai; chua thay chan nham luong nao |")
    d.append("| `SaleStaff` / `SaleManager` | 5 | Ban sao go nham, vo hai vi da co ten dung di kem |")

    out = os.path.join(ROOT, "tests", "L4_ROUTE_DRIFT.md")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(d) + "\n")
    return out


def viet_huong_dan(dem):
    d = f"""# L4 E2E Test — huong dan chay & thuyet trinh

Dot chay 12/08/2026, nhanh `main`, workbook `Report_5_4_L4-E2ETests_VietTien_v1_3.xlsx` (47 case).
Ket qua: **Pass {dem['Pass']} · Fail {dem['Fail']} · Blocked {dem['Blocked']}**.

---

## 1. Chay nhu the nao

**Buoc 0 — cai cong cu (chi 1 lan):**

```bash
cd D:\\SEP_Deploy\\SEP490_fe\\frontend\\e2e && npm ci && npx playwright install chromium
```

Neu `SEP490_fe/frontend/node_modules` chua co thi cai them cho app (dung `npm ci` de KHONG
sua `package-lock.json`):

```bash
cd D:\\SEP_Deploy\\SEP490_fe\\frontend && npm ci
```

**Buoc 1 — bat backend (cua so PowerShell rieng, de chay nen):**

```bash
powershell -ExecutionPolicy Bypass -File D:\\SEP_Deploy\\SEP490_be\\tests\\run-l4-api.ps1
```

**Buoc 2 — chay 47 case (Playwright tu bat Vite dev server):**

```bash
cd D:\\SEP_Deploy\\SEP490_fe\\frontend\\e2e && npx playwright test
```

**Buoc 3 — dien ket qua vao workbook + sinh bao cao:**

```bash
cd D:\\SEP_Deploy\\SEP490_be && python tools\\l4_report.py
```

**Xem bao cao co anh/video/trace cua tung case do:**

```bash
cd D:\\SEP_Deploy\\SEP490_fe\\frontend\\e2e && npx playwright show-report
```

| Buoc | Cong cu | Can server? | Thoi gian | Ket qua xem o |
|---|---|---|---|---|
| 0 | npm + playwright install | Khong | 2 phut | — |
| 1 | run-l4-api.ps1 | (la server) | chay nen | `http://localhost:5050/swagger` |
| 2 | npx playwright test | **Co** | ~3,5 phut | console + `e2e/playwright-report/` |
| 3 | python l4_report.py | Khong | 2 giay | workbook + `tests/L4_*.md` |

---

## 2. Vi sao chay tren DB local chu khong phai Azure

Ke hoach ban dau la chay thang tren Azure SQL that (`viettien.database.windows.net` /
`VietTien2`). **Khong thuc hien duoc**: firewall cua Azure SQL chan IP cua may chay test —

```
Cannot open server 'viettien' requested by the login.
Client with IP address '123.27.14.137' is not allowed to access the server.
```

Backend cung se chet dung loi nay neu tro vao Azure. Vi vay dot nay chay tren
**SQL Server local `VietTien22_L3`**.

Ngoai firewall con mot rui ro thu hai da tranh duoc: `Program.cs:284-291` goi
`db.Database.Migrate()` **vo dieu kien** moi lan khoi dong. Neu ket noi duoc Azure thi chi
viec *bat* backend da lam ALTER schema production truoc khi chay bat ky test nao.

`tests/run-l4-api.ps1` ghi de `ConnectionStrings__DefaultConnection` bang bien moi truong
(thang appsettings trong thu tu cau hinh .NET) nen **khong phai sua file cau hinh nao**.

---

## 3. Bo test nay khong dung vao code ung dung

- `SEP490_fe/frontend/e2e/` co `package.json` + `node_modules` RIENG. `package.json`,
  `package-lock.json` va `vite.config.ts` cua app khong bi sua mot dong nao.
- Khong them `data-testid` vao FE. Selector chi dung `getByRole` / `getByText` / `getByLabel`
  tren DOM san co.
- Khong sua `Program.cs`, `appsettings*.json`, migration, hay `.gitignore` goc.
- Xoa thu muc `e2e/` la repo ve nguyen trang.

---

## 4. Doc ket qua the nao

Bao cao phan biet 3 trang thai, khac nhau ve **y nghia**:

| Trang thai | Nghia | Xu ly |
|---|---|---|
| **Pass** | Chay duoc va dung ky vong workbook | — |
| **Fail** | Co man hinh / co endpoint nhung hanh vi SAI | Phai sua code |
| **Blocked** | Khong co man hinh / endpoint de kiem | Phai lam tinh nang, hoac sua workbook |

{dem['Blocked']} case Blocked KHONG phai la loi cua dot kiem thu — do la phan chuc nang chua ton tai
(trung phan lon voi DEF-L3-004/005 da bao cao o dot L3) hoac workbook ghi sai route.

Chi tiet tung defect: `tests/L4_KET_QUA_TOM_TAT.md`.
Danh sach lech route/endpoint/du lieu: `tests/L4_ROUTE_DRIFT.md`.

---

## 5. Du lieu do test tao ra

Moi ban ghi test tao ra deu mang dau nhan de tim lai:

- Email: `e2e.l4.<case>.<timestamp>@viettien.test`
- Ten / ghi chu: bat dau bang `E2E-L4`
- Don hang: ghi chu `E2E-L4 don do kiem thu tu dong tao`

Moi lenh SQL GHI truc tiep deu duoc luu vao `tests/L4_mutations_2026-08-12.sql` kem thoi diem
chay, de doi chieu ve sau. Trang thai DB truoc dot chay: `tests/L4_baseline_2026-08-12.csv`.
"""
    out = os.path.join(ROOT, "tests", "L4_HUONG_DAN_THUYET_TRINH.md")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(d)
    return out


def main():
    ket_qua = doc_ket_qua()
    print(f"Doc duoc {len(ket_qua)} Test ID tu results.json")
    thong_ke, chua_thay = dien_workbook(ket_qua)
    print(f"Da dien {len(thong_ke)} dong vao workbook")
    if chua_thay:
        print(f"CANH BAO — case trong workbook nhung khong co ket qua chay: {chua_thay}")
    out, dem = viet_tom_tat(thong_ke)
    print(f"Da ghi {out}")
    print(f"Da ghi {viet_route_drift()}")
    print(f"Da ghi {viet_huong_dan(dem)}")
    print(f"Pass={dem['Pass']}  Fail={dem['Fail']}  Blocked={dem['Blocked']}")


if __name__ == "__main__":
    main()
